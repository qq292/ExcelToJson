import openpyxl
import json
import os
import argparse
import re



r'''
一共4个参数( 2个必选参数, 3个可选参数)
第一个参数: 要解析的excel文件
第二个参数: 解析后要保存的路径

[可选参数]
-isAssociation:默认值[1],可选0和1[1=true,0=false],只对dict有效
-targetType:默认值["dict"],可选dict和list
-sub:默认值[None]  正则表达式去掉表格第一行cell值不需要的部分(可以做注释用,或其它用途)

示例
python ExcelToJson.py  "excelPath/xxx.xlsx"  "savePath"  -s "\(.*\)"


'''

class EParse:
    def __init__(self,fileName,sub=None) -> None:
        self.workbook = openpyxl.load_workbook(filename=fileName)
        self.sub=sub
  
    def __rowdata(self,st):
        return st.iter_rows(min_row=2, max_row=st.max_row, min_col=1,max_col=st.max_column)
    
    def __printDict(self,d):
        print( json.dumps(d,ensure_ascii=False,sort_keys=True, indent=4, separators=(',', ':')))
        
    def firstSheetTitle(self):
        return self.workbook.worksheets[0].title
    
    def parseToList(self,st)-> list: 
        return [{self.subs(str(st.cell(row=1,column=col.column).value)):col.value for col in row} for row in self.__rowdata(st)]
         
    def parseToDict(self,st)-> dict: 
        return {str(row[0].value):{self.subs(str(st.cell(row=1,column=col.column).value)):col.value for col in row[1:]} for row in self.__rowdata(st)}
    
    def subs(self,obj):
        if self.sub is None or not isinstance(obj,str):
            return str(obj)
        else:
           return re.sub(self.sub,"",obj)
            
    def parseToDictAssociationSheet(self) -> dict:
        firstSheetDcit = self.parseToDict(self.workbook.worksheets[0])
        for sheet in self.workbook.worksheets[1:]:
            key = sheet.title
            for k,v in self.parseToDict(sheet).items():
                if k in firstSheetDcit:
                    firstSheetDcit[k][key]=v
            
        return firstSheetDcit
        


def parse_args():
    parser = argparse.ArgumentParser(
        description="...->--")
    parser.add_argument('excelPath', help="excelPath")
    parser.add_argument('savePath', default=os.getcwd(), help="SavePath")
    parser.add_argument('-a', default=1, type=int, help="isAssociation")
    parser.add_argument('-t', default="dict", type=str, help="targetType")
    parser.add_argument('-s', default=None, type=str, help="sub")
    args = parser.parse_args()
    return args

def SaveToJson(data:dict,path:str):
    with open(path, "w",encoding='utf-8') as f:
        f.write(data)
    print("已保存到:  "+path)



if __name__ == "__main__":
    args = parse_args()
    excelPath = args.excelPath
    savePath = args.savePath
    targetType = args.t
    isAssociation = args.a
    sub = args.s
    ep=EParse(excelPath,sub)
    if targetType.title().__eq__("Dict"):
        if isAssociation!=0:
            result = ep.parseToDictAssociationSheet()
            path=f"{savePath}{os.sep}{ep.firstSheetTitle()}.json"
            data=json.dumps(result,ensure_ascii=False,sort_keys=True, indent=4, separators=(',', ':'))
            SaveToJson(data,path)
        else:
            for sheet in ep.workbook.worksheets:
                path=f"{savePath}{os.sep}{sheet.title}.json"
                data=json.dumps(ep.parseToDict(sheet),ensure_ascii=False,sort_keys=True, indent=4, separators=(',', ':'))
                SaveToJson(data,path)
    elif targetType.title().__eq__("List"):
         for sheet in ep.workbook.worksheets:
                path=f"{savePath}{os.sep}{sheet.title}.json"
                data=json.dumps(ep.parseToList(sheet),ensure_ascii=False,sort_keys=True, indent=4, separators=(',', ':')) 
                SaveToJson(data,path)
        
        
        